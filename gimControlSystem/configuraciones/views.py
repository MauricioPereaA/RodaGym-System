from django.shortcuts import render, redirect
from django.views.generic import TemplateView
from django.views import View
from django.utils import timezone
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test
from django.utils.decorators import method_decorator
from finanzas.models import Transaccion, Ticket, ItemTicket
from miembros.models import Miembro
from acceso.models import ConfiguracionesAcesso
from inventario.models import Producto
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from django.core.serializers.json import DjangoJSONEncoder
from django.contrib.auth.decorators import login_required
import json
from escpos import *
from django.conf import settings
import os
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@method_decorator(user_passes_test(lambda u: u.is_staff), name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class ConfiguracionesView(View):
    template_name = 'configuraciones/configuraciones.html'

    def get(self, request, *args, **kwargs):
        configuracion = ConfiguracionesAcesso.objects.first()  # Asumiendo que solo tienes una configuración global
        context = {'configuracion': configuracion}
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        configuracion = ConfiguracionesAcesso.objects.first()  # Asumiendo que solo tienes una configuración global

        if request.POST.get('mensaje_global') == 'on':
            mensaje_global = True
            mensaje_global_str = request.POST.get('mensaje_global_str')
        else:
            mensaje_global = False
            mensaje_global_str = configuracion.mensaje_global_str

        configuracion.mensaje_global = mensaje_global
        configuracion.mensaje_global_str = mensaje_global_str
        configuracion.save()

        return redirect('configuraciones:configuraciones')

@method_decorator(user_passes_test(lambda u: u.is_staff), name='dispatch')
class ReportesView(TemplateView):
    template_name = 'configuraciones/generacion_reportes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuarios'] = User.objects.all()
        context['transacciones'] = []
        context['miembros'] = []
        context['inventario'] = []
        return context

    def post(self, request, *args, **kwargs):
        context = self.get_context_data()
        fecha_desde = request.POST.get('fecha_desde')
        fecha_hasta = request.POST.get('fecha_hasta')
        filtro_principal = request.POST.get('filtro_principal')
        tipo_transaccion = request.POST.get('tipo_transaccion')

        if fecha_desde and fecha_hasta:
            try:
                fecha_desde = timezone.datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                fecha_hasta = timezone.datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            except ValueError:
                context['error'] = 'Formato de fecha incorrecto.'

            if filtro_principal == 'transacciones':
                transacciones = Transaccion.objects.filter(fecha__range=(fecha_desde, fecha_hasta))
                if tipo_transaccion == 'ventas_general':
                    transacciones = transacciones.filter(venta_general=True)
                elif tipo_transaccion == 'membresias':
                    transacciones = transacciones.filter(venta_general=False)

                transacciones_dict = [
                    {
                        'id': transaccion.id,
                        'fecha': transaccion.fecha.isoformat(),
                        'usuario': transaccion.usuario.username,
                        'cantidad': str(transaccion.cantidad),
                        'descripcion': transaccion.descripcion,
                        'metodo_pago': transaccion.metodo_pago
                    }
                    for transaccion in transacciones
                ]
                request.session['transacciones'] = json.dumps(transacciones_dict)
                context['transacciones'] = transacciones_dict

            elif filtro_principal == 'inscripciones':
                miembros = Miembro.objects.filter(fecha_inicio_membresia__range=(fecha_desde, fecha_hasta))
                miembros_dict = [
                    {
                        'id': miembro.id,
                        'nombres': miembro.nombres,
                        'apellidos': miembro.apellidos,
                        'fecha_inicio_membresia': miembro.fecha_inicio_membresia.isoformat()
                    }
                    for miembro in miembros
                ]
                request.session['miembros'] = json.dumps(miembros_dict)
                context['miembros'] = miembros_dict

            elif filtro_principal == 'inventario':
                inventario = Producto.objects.all()
                inventario_dict = [
                    {
                        'id': item.id,
                        'nombre': item.nombre,
                        'codigo_barras': item.codigo_barras,
                        'precio': str(item.precio),
                        'total_bodega': item.total_bodega,
                        'tipo': item.tipo
                    }
                    for item in inventario
                ]
                request.session['inventario'] = json.dumps(inventario_dict)
                context['inventario'] = inventario_dict

        if 'exportar_excel' in request.POST:
            transacciones = json.loads(request.session.get('transacciones', '[]'))
            miembros = json.loads(request.session.get('miembros', '[]'))
            inventario = json.loads(request.session.get('inventario', '[]'))
            if transacciones:
                response = self.exportar_excel(transacciones=transacciones)
            elif miembros:
                response = self.exportar_excel(miembros=miembros)
            elif inventario:
                response = self.exportar_excel(inventario=inventario)

            # Limpiar la sesión después de la exportación
            if 'transacciones' in request.session:
                del request.session['transacciones']
            if 'miembros' in request.session:
                del request.session['miembros']
            if 'inventario' in request.session:
                del request.session['inventario']

            return response

        return render(request, self.template_name, context)

    def exportar_excel(self, transacciones=None, miembros=None, inventario=None):
        wb = Workbook()
        name_book = ''
        if transacciones:
            name_book = 'attachment; filename=transacciones_reporte.xlsx'
            ws1 = wb.active
            ws1.title = "Transacciones"
            # Header
            headers_transacciones = ["ID", "Fecha", "Usuario", "Cantidad", "Descripción", "MetodoPago"]
            ws1.append(headers_transacciones)
            for col in range(1, len(headers_transacciones) + 1):
                ws1.cell(row=1, column=col).font = Font(bold=True)

            for transaccion in transacciones:
                ws1.append([transaccion['id'], transaccion['fecha'], transaccion['usuario'], transaccion['cantidad'], transaccion['descripcion'], transaccion['metodo_pago']])
        elif miembros:
            name_book = 'attachment; filename=miembros.xlsx'
            ws1 = wb.active
            ws1.title = "Inscripciones"
            # Header
            headers_miembros = ["ID", "Nombre", "Apellido", "Fecha de Inicio"]
            ws1.append(headers_miembros)
            for col in range(1, len(headers_miembros) + 1):
                ws1.cell(row=1, column=col).font = Font(bold=True)

            for miembro in miembros:
                ws1.append([miembro['id'], miembro['nombres'], miembro['apellidos'], miembro['fecha_inicio_membresia']])
        elif inventario:
            name_book = 'attachment; filename=inventario.xlsx'
            ws1 = wb.active
            ws1.title = "Inventario Actual"
            # Header
            headers_inventario = ["ID", "Nombre", "Código de Barras", "Precio", "Total Bodega", "Tipo"]
            ws1.append(headers_inventario)
            for col in range(1, len(headers_inventario) + 1):
                ws1.cell(row=1, column=col).font = Font(bold=True)

            for item in inventario:
                ws1.append([item['id'], item['nombre'], item['codigo_barras'], item['precio'], item['total_bodega'], item['tipo']])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = name_book
        wb.save(response)
        return response


@method_decorator(user_passes_test(lambda u: u.is_staff), name='dispatch')
class AltaUsuarioView(TemplateView):
    template_name = 'configuraciones/alta_usuario.html'

    def post(self, request, *args, **kwargs):
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email')
        
        if username and password and email:
            user = User.objects.create_user(username=username, email=email, password=password)
            user.save()
            return render(request, self.template_name, {'success': 'Usuario creado correctamente'})
        return render(request, self.template_name, {'error': 'Faltan datos'})

@login_required
@user_passes_test(lambda u: u.is_staff)
def reimpresion_ticket_view(request):
    if request.method == 'POST':
        ticket_id = request.POST.get('ticket_id')
        try:
            ticket = Ticket.objects.get(id=ticket_id)
            items_ticket = ItemTicket.objects.filter(ticket=ticket)

            # Reimpresion ticket
            ruta_imagen = os.path.join(settings.BASE_DIR, 'static', 'images', 'spacefit', 'logo_impresion.png')
            thermal_printer = printer.Win32Raw('POS58 Printer')  # Cambia esto según tu configuración de impresora
            thermal_printer.set_with_default()
            thermal_printer.set(align='center')
            thermal_printer.image(ruta_imagen)
            thermal_printer.textln('Avenida Elias Zamora, Los Mangos')
            thermal_printer.textln('28869-Manzanillo, Colima')
            thermal_printer.textln('================================')
            thermal_printer.set(align='left')
            thermal_printer.textln(f'FOLIO:{ticket.id}')
            ticket_fecha_str = ticket.fecha.strftime('%Y-%m-%d %H:%M')
            thermal_printer.textln(f'FECHA:{ticket_fecha_str}')
            thermal_printer.textln(f'CAJERO:{ticket.cajero}')
            thermal_printer.textln(f'MIEMBROID:{ticket.miembro.nombres}')
            thermal_printer.textln('================================')
            thermal_printer.set(align='center')
            thermal_printer.set(bold=True)
            thermal_printer.textln('CANT   DESCRIPCION       IMPORTE')
            thermal_printer.textln('================================')
            thermal_printer.set(bold=False)
            for item in items_ticket:
                itemCantidad = str(item.cantidad)
                itemDescripcion = item.descripcion[:19]
                itemImporte = str(item.importe)
                thermal_printer.textln(f'{itemCantidad:<4}{itemDescripcion:<19} ${itemImporte:>7}')
            thermal_printer.textln('________________________________')
            thermal_printer.textln(f'                 TOTAL: ${ticket.total:>7}')
            thermal_printer.textln('================================')
            thermal_printer.set(align='right')
            thermal_printer.textln(f'RECIBIDO: ${ticket.recibido:>7}')
            thermal_printer.textln(f'CAMBIO: ${ticket.cambio:>7}')
            thermal_printer.textln(f'FORMA PAGO:{ticket.metodo_pago}')
            thermal_printer.textln('================================')
            thermal_printer.ln(2)
            thermal_printer.close()

            return JsonResponse({'success': True})
        except Ticket.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Ticket no encontrado'})
    return render(request, 'configuraciones/reimprimir_ticket.html')

